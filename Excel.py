import pandas as pd
from openpyxl import Workbook
import openpyxl,ipaddress,netaddr

class Excel:

#Création d'une liste pour stocker les élémenents du fichier xlsx
    def __init__(self):
        self.files = []
        self.grid = []
        self.dico = {None:-1,'00:00:00:00:00:00':0}
        self.compteur = 1
        self.new_file = [] 
        

#Ajoute un fichier a la liste
    def add_file(self,new_file):
        self.files.append(new_file)


#Change les élements de la liste liste en entier
    def change_to_int_list(self,liste):
        new_liste =[]
        for i in liste:
            try:
                col = float(i)
                new_liste.append(col)
            except:
                try:
                    ip = int(netaddr.IPAddress(i))
                    new_liste.append(str(ip))
                except:
                    if i not in self.dico:
                        self.dico[i] = self.compteur
                        self.compteur+=1
                    new_liste.append(self.dico[i])
        self.new_file.append(new_liste)

#Récupération de tout les elments du fichier xlsx
    def read_clean_file(self,name,maxi=-1):
        print('Reading file ...')
        book = openpyxl.load_workbook(name,read_only=True)
        sheet = book.active
        book_clean = Workbook()
        sheet_clean = book_clean.active
        print('File readed')
        c=0
        for value in sheet.iter_rows(min_row=2, min_col=2, values_only=True): 
            c+=1
            self.grid.append(value)
            if c%5000 == 0:
                print("Dans le for",c)
            if maxi > 0 and c > maxi:
                break
            self.change_to_int_list(self.grid[0])
            sheet_clean.append(self.new_file[0])
            self.grid = []
            self.new_file = []
            print("end")
        book_clean.save("test3.xlsx")
